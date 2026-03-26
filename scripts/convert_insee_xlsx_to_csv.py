"""
Conversion du fichier Excel INSEE en CSV pour Snowflake.

Source : estim-pop-nreg-sexe-aq-1975-2026.xlsx
         https://www.insee.fr/fr/statistiques/8721456
         (Estimations de population par région, sexe et âge, 1975-2026)

Produit : insee_population_region.csv
          Format plat : REGION | SEXE | AGE_GROUP | POPULATION | ANNEE

Le fichier Excel INSEE est inexploitable tel quel :
52 onglets, en-têtes sur 4 lignes, 64 colonnes, cellules fusionnées.
Ce script extrait les 4 années utiles (2022-2025) et aplatit le tout
en un CSV propre chargeable dans Snowflake via Load Data.

Nécessite : pip install openpyxl
"""

import openpyxl
import csv

# --- Config ---

FICHIER_INSEE = "estim-pop-nreg-sexe-aq-1975-2026.xlsx"
FICHIER_CSV = "insee_population_region.csv"
ANNEES = [2022, 2023, 2024, 2025]

# 20 tranches quinquennales, dans l'ordre des colonnes Excel
TRANCHES_AGE = [
    '0 à 4 ans', '5 à 9 ans', '10 à 14 ans', '15 à 19 ans',
    '20 à 24 ans', '25 à 29 ans', '30 à 34 ans', '35 à 39 ans',
    '40 à 44 ans', '45 à 49 ans', '50 à 54 ans', '55 à 59 ans',
    '60 à 64 ans', '65 à 69 ans', '70 à 74 ans', '75 à 79 ans',
    '80 à 84 ans', '85 à 89 ans', '90 à 94 ans', '95 ans et plus'
]

# Structure de chaque onglet : Régions | Ensemble (20 col) | Total | Hommes (20 col) | Total | Femmes (20 col) | Total
# L'offset indique la première colonne de données pour chaque sexe
SEXES = {
    'Ensemble': 1,
    'Hommes': 22,
    'Femmes': 43
}

# Lignes à ignorer (totaux nationaux, notes de bas de page)
REGIONS_A_EXCLURE = {'France métropolitaine', 'France', 'France métropolitaine et DOM', None}
PREFIXES_A_IGNORER = ('Source', 'Champ', 'Note')

# --- Extraction ---

print(f"Ouverture de {FICHIER_INSEE}...")
wb = openpyxl.load_workbook(FICHIER_INSEE, read_only=True)
print(f"{len(wb.sheetnames)} onglets, extraction de {ANNEES}")

lignes_csv = []

for annee in ANNEES:
    print(f"\n{annee}...")
    ws = wb[str(annee)]
    nb_regions = 0

    # Données à partir de la ligne 6 (1-4 = en-têtes, 5 = noms de colonnes)
    for row in ws.iter_rows(min_row=6, values_only=True):
        region = row[0]

        if not region or region in REGIONS_A_EXCLURE:
            continue
        if any(region.startswith(p) for p in PREFIXES_A_IGNORER):
            continue

        nb_regions += 1

        for sexe, offset in SEXES.items():
            for i, tranche in enumerate(TRANCHES_AGE):
                valeur = row[offset + i]
                if valeur is not None:
                    lignes_csv.append([region, sexe, tranche, int(valeur), annee])

    print(f"  {nb_regions} régions")

# --- Export ---

print(f"\nÉcriture de {FICHIER_CSV}...")
with open(FICHIER_CSV, 'w', newline='', encoding='utf-8') as f:
    writer = csv.writer(f)
    writer.writerow(['REGION', 'SEXE', 'AGE_GROUP', 'POPULATION', 'ANNEE'])
    writer.writerows(lignes_csv)

# --- Vérif rapide ---

print(f"\n{len(lignes_csv)} lignes exportées")
print(f"Régions : {sorted(set(r[0] for r in lignes_csv))}")
print(f"Sexes : {sorted(set(r[1] for r in lignes_csv))}")
print(f"Années : {sorted(set(r[4] for r in lignes_csv))}")
print(f"\n→ Charger dans OC_P8.RAW_INSEE.POPULATION_REGION via Snowflake Load Data")